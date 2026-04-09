from datetime import date, datetime
from django.db.models import Q, TextField, Value
from django.http import JsonResponse
from django.urls import reverse
from django.shortcuts import render
from autorisations.models.models_instruction import Dossier, DossierManifSportive, DossierManifestationLiaison, EtapeDossier, Demarche
from autorisations.models.models_avis import Expert
from autorisations.models.models_utilisateurs import ContactExterne, DossierBeneficiaire, DossierInterlocuteur, Groupeinstructeur, Instructeur
from django.contrib.auth.decorators import login_required
from django.db.models.functions import ExtractYear, Coalesce
from django.db.models import Value, F
from django.db.models.functions import Concat
from autorisations.models.models_avis import Avis
from django.db.models.functions import ExtractYear, ExtractMonth

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def clean_int(value):
    return value if value and value.isdigit() else None


# Export Excel Dossiers
# def _export_dossiers_xlsx(dossiers_qs):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Dossiers"

#     headers = [
#         "N°", "Nom dossier", "Démarche", "Étape",
#         "Date dépôt", "Groupe instructeur", "Bénéficiaire"
#     ]
#     ws.append(headers)

#     # Style en-têtes
#     bold = Font(bold=True)
#     for col_idx in range(1, len(headers) + 1):
#         cell = ws.cell(row=1, column=col_idx)
#         cell.font = bold
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

#     # Remplissage
#     for d in dossiers_qs:
#         # Bénéficiaire : on recopie la logique du template :contentReference[oaicite:4]{index=4}
#         beneficiaire = ""
#         interlocuteur = d.dossierinterlocuteur_set.first()
#         if interlocuteur:
#             db = interlocuteur.dossierbeneficiaire_set.select_related("id_beneficiaire").first()
#             if db and db.id_beneficiaire:
#                 b = db.id_beneficiaire
#                 if getattr(b, "raison_sociale", None):
#                     beneficiaire = b.raison_sociale
#                 elif getattr(b, "nom", None) and getattr(b, "prenom", None):
#                     beneficiaire = f"{b.nom} {b.prenom}"

#         nom_dossier = getattr(d, "nom_dossier_plus_parlant", None) or d.nom_dossier

#         ws.append([
#             d.numero,
#             nom_dossier,
#             d.id_demarche.type if d.id_demarche else "",
#             d.id_etape_dossier.etape if d.id_etape_dossier else "",
#             d.date_depot.strftime("%d/%m/%Y") if d.date_depot else "",
#             d.id_groupeinstructeur.nom if d.id_groupeinstructeur else "",
#             beneficiaire,
#         ])

#     # Largeurs de colonnes (simple)
#     widths = [10, 50, 35, 25, 14, 30, 35, 25]
#     for i, w in enumerate(widths, start=1):
#         ws.column_dimensions[get_column_letter(i)].width = w

#     # Output
#     bio = BytesIO()
#     wb.save(bio)
#     bio.seek(0)

#     response = HttpResponse(
#         bio.getvalue(),
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = 'attachment; filename="dossiers.xlsx"'
#     return response


def _export_dossiers_xlsx(resultats):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dossiers"

    headers = [
        "N°", "Nom dossier", "Démarche", "Étape",
        "Date dépôt", "Groupe instructeur", "Bénéficiaire"
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for d in resultats:
        ws.append([
            d["numero"],
            d["nom_dossier"],
            d["demarche"],
            d["etape"],
            d["date_depot"].strftime("%d/%m/%Y") if d["date_depot"] else "",
            d["groupe"],
            d["beneficiaire"],
        ])

    widths = [10, 50, 35, 25, 14, 30, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="dossiers.xlsx"'
    return response






# Export Excel Avis
def _export_avis_xlsx(avis_iterable):
    wb = Workbook()
    ws = wb.active
    ws.title = "Avis"

    headers = [
        "N° Avis",
        "Date demande",
        "Démarche",
        "Expert",
        "Demandeur",
        "Réponse",
        "Publié au RAA"
    ]
    ws.append(headers)

    # Style entêtes
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for a in avis_iterable:
        # Date
        date_demande = ""
        if getattr(a, "date_demande_avis", None):
            date_demande = a.date_demande_avis.strftime("%d/%m/%Y")

        # Démarche
        demarche = ""
        dossier = getattr(a, "id_dossier", None)
        if dossier and getattr(dossier, "id_demarche", None):
            demarche = dossier.id_demarche.type or ""

        # Expert / demandeur (dans ton tableau: a.id_expert, a.id_instructeur)
        expert = str(getattr(a, "id_expert", "") or "")
        demandeur = str(getattr(a, "id_instructeur", "") or "")

        # Réponse (copie exacte de ta logique template)
        favorable = getattr(a, "favorable", None)
        sous_reserve = getattr(a, "sous_reserve", None)

        if favorable is True and not sous_reserve:
            reponse = "Favorable"
        elif favorable is True and sous_reserve:
            reponse = "Favorable sous réserve"
        elif favorable is False:
            reponse = "Défavorable"
        else:
            reponse = "En attente"

        publie = "Oui" if getattr(a, "publie_au_raa", False) else "Non"

        ws.append([
            a.id,
            date_demande,
            demarche,
            expert,
            demandeur,
            reponse,
            publie,
        ])

    # Largeurs de colonnes (simple + lisible)
    widths = [10, 14, 35, 25, 25, 22, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="avis.xlsx"'
    return response



@login_required
def requete_dossiers(request):
    dossiers = Dossier.objects.all().select_related("id_demarche", "id_groupeinstructeur", "id_etape_dossier")
    
    # Dossiers DM non liés
    dossiers_deja_lies_ids = DossierManifestationLiaison.objects.values_list("id_dossier_manif_id", flat=True)
    dossiers_dm = (
        DossierManifSportive.objects
        .exclude(id__in=dossiers_deja_lies_ids)
        .select_related("id_etape")
    )

    # POUR LE MOMENT ON EXCLU MANIFESTATIONS SPORTIVES
    # dossiers = dossiers.exclude(id_demarche__type="Manifestations sportives")


    # Menus déroulant
    etapes_dossier = EtapeDossier.objects.all().order_by('etape')
    demarches = Demarche.objects.all().order_by('type')
    groupes = Groupeinstructeur.objects.all().order_by('nom')

    annees_dossiers = set(Dossier.objects.annotate(annee=ExtractYear("date_depot")).values_list("annee", flat=True))
    annees_dm = set(DossierManifSportive.objects.annotate(annee=ExtractYear("date_depot")).values_list("annee", flat=True))
    annees = sorted({a for a in (annees_dossiers | annees_dm) if a is not None}, reverse=True)
    # annees = Dossier.objects.annotate(annee=ExtractYear("date_depot")) \
    #                     .values_list("annee", flat=True).distinct().order_by("-annee")
    
    annees_avis = Avis.objects.annotate(annee=ExtractYear("date_demande_avis")) \
        .values_list("annee", flat=True).distinct().order_by("-annee")
    mois_list = Avis.objects.annotate(mois=ExtractMonth("date_demande_avis")) \
        .values_list("mois", flat=True).distinct().order_by("mois")
    

    numero = clean_int(request.GET.get("numero"))
    # date_depot = request.GET.get("date_depot")
    date_debut_reception = request.GET.get("date_debut_reception")
    date_fin_reception = request.GET.get("date_fin_reception")
    date_debut_instruction = request.GET.get("date_debut_instruction")
    date_fin_instruction = request.GET.get("date_fin_instruction")

    type_demarche = request.GET.get("d_type_demarche")
    groupe = request.GET.get("groupe")
    etape = request.GET.get("etape")
    nom = request.GET.get("nom")
    annee = request.GET.get("d_annee")
    instructeur = request.GET.get("instructeur")

    if annee :
        dossiers = dossiers.filter(date_depot__year=annee)
        dossiers_dm = dossiers_dm.filter(date_depot__year=annee)

    if numero:
        dossiers = dossiers.filter(numero=numero)
        dossiers_dm = dossiers_dm.filter(numero_dossier_declaration_manifestations=numero)

    # Période ou le dossier a été recu
    # Si une date de début est fournie sans date de fin → aujourd'hui
    if date_debut_reception and not date_fin_reception:
        date_fin_reception = date.today().isoformat()

    # Si date debut > date fin : on inverse les dates
    if date_debut_reception and date_fin_reception and date_debut_reception > date_fin_reception :
        date_debut_reception, date_fin_reception = (date_fin_reception, date_debut_reception)

    if date_debut_reception:
        dossiers = dossiers.filter(date_depot__date__gte=date_debut_reception)
        dossiers_dm = dossiers_dm.filter(date_depot__date__gte=date_debut_reception)

    if date_fin_reception:
        dossiers = dossiers.filter(date_depot__date__lte=date_fin_reception)
        dossiers_dm = dossiers_dm.filter(date_depot__date__lte=date_fin_reception)


    # Période ou le dossier a été instruit
    if date_debut_instruction and not date_fin_instruction:
        date_fin_instruction = date.today().isoformat()

    if date_debut_instruction and date_fin_instruction:
        if date_debut_instruction > date_fin_instruction:
            date_debut_instruction, date_fin_instruction = (
                date_fin_instruction, date_debut_instruction
            )

    if date_debut_instruction:
        dossiers = dossiers.filter(date_fin_instruction__date__gte=date_debut_instruction)
        dossiers_dm = dossiers_dm.filter(avis__date_reponse__date__gte=date_debut_instruction)

    if date_fin_instruction:
        dossiers = dossiers.filter(date_fin_instruction__date__lte=date_fin_instruction)
        dossiers_dm = dossiers_dm.filter(avis__date_reponse__date__lte=date_fin_instruction)



    if type_demarche :
        dossiers = dossiers.filter(id_demarche__type__icontains=type_demarche)

        if "manifestations sportives".find(type_demarche.lower()) == -1:
            dossiers_dm = dossiers_dm.none()

    if groupe :
        dossiers = dossiers.filter(id_groupeinstructeur__nom__icontains=groupe)
        if "manifestations sportives".find(groupe.lower()) == -1:
            dossiers_dm = dossiers_dm.none()

    if etape :
        dossiers = dossiers.filter(id_etape_dossier__etape__icontains=etape)
        dossiers_dm = dossiers_dm.filter(id_etape__etape__icontains=etape)

    if instructeur:
        exists_instructeur = Instructeur.objects.annotate(
            nom_complet=Concat(
                F("id_agent_autorisations__nom"),
                Value(" "),
                F("id_agent_autorisations__prenom"),
            )
        ).filter(nom_complet__iexact=instructeur).exists()
        
        if exists_instructeur :
            dossiers = dossiers.annotate(
                nom_complet_instructeur=Concat(
                    F("dossierinstructeur__id_instructeur__id_agent_autorisations__nom"),
                    Value(" "),
                    F("dossierinstructeur__id_instructeur__id_agent_autorisations__prenom")
                )
            ).filter(
                Q(nom_complet_instructeur__iexact=instructeur)
            ).distinct()

        else:
            instructeur = ""

        # Dossiers DM non liés
        dossiers_dm = dossiers_dm.none()

    if nom:
        # Vérifier si le nom existe en tant que raison sociale
        exists_raison = ContactExterne.objects.filter(
            dossierbeneficiaire__isnull=False,
            raison_sociale__iexact=nom
        ).exists()

        # Vérifier si le nom existe comme Nom Prénom
        exists_nom_prenom = ContactExterne.objects.annotate(
            nom_complet=Concat(
                F("nom"),
                Value(" "),
                F("prenom")
            )
        ).filter(
            dossierbeneficiaire__isnull=False,
            nom_complet__iexact=nom
        ).exists()

        nom_valide = exists_raison or exists_nom_prenom
        if nom_valide:
            dossiers = dossiers.annotate(
                nom_complet_beneficiaire=Concat(
                    F("dossierinterlocuteur__dossierbeneficiaire__id_beneficiaire__nom"),
                    Value(" "),
                    F("dossierinterlocuteur__dossierbeneficiaire__id_beneficiaire__prenom")
                )
            ).filter(
                Q(dossierinterlocuteur__dossierbeneficiaire__id_beneficiaire__raison_sociale__iexact=nom) |
                Q(nom_complet_beneficiaire__iexact=nom)
            ).distinct()

        else:
            dossiers = dossiers.none()

        # Dossiers DM non liés
        dossiers_dm = dossiers_dm.annotate(
            nom_prenom=Concat(
                F("nom_organisateur"),
                Value(" "),
                F("prenom_organisateur"),
                output_field=TextField()
            ),
            prenom_nom=Concat(
                F("prenom_organisateur"),
                Value(" "),
                F("nom_organisateur"),
                output_field=TextField()
            )
        ).filter(
            Q(nom_prenom__icontains=nom) |
            Q(prenom_nom__icontains=nom) |
            Q(nom_organisateur__icontains=nom) |
            Q(prenom_organisateur__icontains=nom) |
            Q(structure__icontains=nom)
        )
        
 
    dossiers = dossiers.order_by("-date_depot").distinct()


    # 1) On prend les ids uniques pour éviter des doublons
    ids = dossiers.values("id").distinct()
    dossiers = (
        Dossier.objects
        .filter(id__in=ids)
        .select_related("id_demarche", "id_groupeinstructeur", "id_etape_dossier")
        .order_by("-date_depot")
    )

    dossiers_dm = dossiers_dm.distinct()


    resultats_dossiers = []


    # # Ajout du champ 'lien' pour chaque dossier
    # for d in dossiers:
    #     if d.id_etape_dossier and d.id_etape_dossier.etape == "À affecter":
    #         d.lien = f"/preinstruction/{d.numero}/"
    #     else:
    #         d.lien = f"/instruction/{d.numero}/"


    # DOSSIERS CLASSIQUES
    for d in dossiers:
        beneficiaire = ""
        interlocuteur = d.dossierinterlocuteur_set.first()
        if interlocuteur:
            db = interlocuteur.dossierbeneficiaire_set.select_related("id_beneficiaire").first()
            if db and db.id_beneficiaire:
                b = db.id_beneficiaire
                if getattr(b, "raison_sociale", None):
                    beneficiaire = b.raison_sociale
                elif getattr(b, "nom", None) and getattr(b, "prenom", None):
                    beneficiaire = f"{b.nom} {b.prenom}"

        resultats_dossiers.append({
            "source": "dossier",
            "obj": d,
            "numero": d.numero,
            "nom_dossier": getattr(d, "nom_dossier_plus_parlant", None) or d.nom_dossier,
            "demarche": d.id_demarche.type if d.id_demarche else "",
            "etape": d.id_etape_dossier.etape if d.id_etape_dossier else "",
            "date_depot": d.date_depot,
            "groupe": d.id_groupeinstructeur.nom if d.id_groupeinstructeur else "",
            "beneficiaire": beneficiaire,
            "lien": reverse("preinstruction_dossier", kwargs={"numero": d.numero})
                    if d.id_etape_dossier and d.id_etape_dossier.etape == "À affecter"
                    else reverse("instruction_dossier", kwargs={"num_dossier": d.numero}),
        })


    # DOSSIERS DM
    groupe_manif = Groupeinstructeur.objects.filter(nom__iexact="Manifestations sportives").first()
    for d in dossiers_dm.order_by("-date_depot"):
        nom_complet = " ".join(
            part.strip()
            for part in [d.prenom_organisateur, d.nom_organisateur]
            if part and part.strip()
        )

        structure = d.structure.strip() if d.structure and d.structure.strip() else ""
        beneficiaire = nom_complet or structure or ""
        print(beneficiaire)

        url_name = ("dossier_manif_sportive_sans_ds_archive" if d.archive else "dossier_manif_sportive_sans_ds")

        resultats_dossiers.append({
            "source": "dossier_dm",
            "obj": d,
            "numero": d.numero_dossier_declaration_manifestations,
            "nom_dossier": d.nom_dossier or "",
            "demarche": "Manifestations sportives",
            "etape": d.id_etape.etape if d.id_etape else "",
            "date_depot": d.date_depot,
            "groupe": groupe_manif.nom if groupe_manif else "Manifestations sportives",
            "beneficiaire": beneficiaire,
            "lien": reverse(url_name, kwargs={"numero": d.numero_dossier_declaration_manifestations}),
        })

    resultats_dossiers.sort(
        key=lambda x: x["date_depot"] or datetime.min,
        reverse=True
    )


    # Export Excel si demandé
    # if request.GET.get("export") == "xlsx":
    #     # Optionnel mais mieux: limiter les requêtes bénéficiaire
    #     dossiers = dossiers.prefetch_related("dossierinterlocuteur_set__dossierbeneficiaire_set__id_beneficiaire")
    #     return _export_dossiers_xlsx(dossiers)

    if request.GET.get("export") == "xlsx":
        return _export_dossiers_xlsx(resultats_dossiers)
            

    context = {
        "dossiers": resultats_dossiers,
        "etapes": EtapeDossier.objects.all(),
        "demarches": Demarche.objects.all(),
        'etapes_dossier': etapes_dossier,
        "demarches": demarches,
        "groupes": groupes,
        "recherche_dossier_effectuee": bool(request.GET),
        "recherche_avis_effectuee": False,
        "annees": annees,
        "annees_avis": annees_avis,
        "mois_list": mois_list,


        # Champs nettoyés pour pré-remplissage propre
        "nom_rempli": nom or "",
        "instructeur_rempli": instructeur or "",
        "numero_rempli": numero or "",
        "date_debut_reception_rempli": date_debut_reception or "",
        "date_fin_reception_rempli": date_fin_reception or "",
        "date_debut_instruction_rempli": date_debut_instruction or "",
        "date_fin_instruction_rempli": date_fin_instruction or "",
    }
    return render(request, "instruction/requetes.html", context)




@login_required
def requete_avis(request):

    # --- Base queryset ---
    avis_list = Avis.objects.select_related(
        "id_dossier",
        "id_demarche",
        "id_instructeur",
        "id_expert",
    ).filter(statut="Envoyé")


    # Menus déroulants ---
    etapes_dossier = EtapeDossier.objects.all().order_by('etape')
    demarches = Demarche.objects.all().order_by('type')
    groupes = Groupeinstructeur.objects.all().order_by('nom')
    annees = Dossier.objects.annotate(annee=ExtractYear("date_depot")) \
                        .values_list("annee", flat=True).distinct().order_by("-annee")
    
    annees_avis = Avis.objects.annotate(annee=ExtractYear("date_demande_avis")) \
        .values_list("annee", flat=True).distinct().order_by("-annee")
    mois_list = Avis.objects.annotate(mois=ExtractMonth("date_demande_avis")) \
        .values_list("mois", flat=True).distinct().order_by("mois")

    date_debut_demande_avis = request.GET.get("date_debut_demande_avis")
    date_fin_demande_avis = request.GET.get("date_fin_demande_avis")

    demarches = Demarche.objects.all().order_by("type")


    # --- Récupération des filtres GET ---
    num_avis = clean_int(request.GET.get("num_avis"))
    mois = request.GET.get("mois")
    annee = request.GET.get("a_annee")
    reponse = request.GET.get("reponse")
    demandeur = request.GET.get("demandeur")
    expert = request.GET.get("expert")
    num_dossier = clean_int(request.GET.get("num_dossier"))
    type_demarche = request.GET.get("a_type_demarche")
    publie_raa = request.GET.get("publie_raa")  # Nouveau filtre

    # --- Application des filtres ---
    if num_avis:
        avis_list = avis_list.filter(id=num_avis)

    if mois:
        avis_list = avis_list.filter(date_demande_avis__month=mois)

    if annee:
        avis_list = avis_list.filter(date_demande_avis__year=annee)

    # Période ou l'avis a été demandé
    # Si une date de début est fournie sans date de fin → aujourd'hui
    if date_debut_demande_avis and not date_fin_demande_avis:
        date_fin_demande_avis = date.today().isoformat()

    # Si date debut > date fin : on inverse les dates
    if date_debut_demande_avis and date_fin_demande_avis:
        if date_debut_demande_avis > date_fin_demande_avis:
            date_debut_demande_avis, date_fin_demande_avis = (
                date_fin_demande_avis, date_debut_demande_avis
            )

    if date_debut_demande_avis:
        avis_list = avis_list.filter(
            date_demande_avis__date__gte=date_debut_demande_avis
        )

    if date_fin_demande_avis:
        avis_list = avis_list.filter(
            date_demande_avis__date__lte=date_fin_demande_avis
        )

    if reponse:
        reponse = reponse.strip()
        if reponse == "En attente":
            avis_list = avis_list.filter(favorable__isnull=True)
        elif reponse == "Favorable":
            avis_list = avis_list.filter(favorable=True, sous_reserve=False)
        elif reponse == "Favorable sous réserve":
            avis_list = avis_list.filter(favorable=True, sous_reserve=True)
        elif reponse == "Défavorable":
            avis_list = avis_list.filter(favorable=False)

    if publie_raa:
        if publie_raa.lower() == "oui":
            avis_list = avis_list.filter(publie_au_raa=True)
        elif publie_raa.lower() == "non":
            avis_list = avis_list.filter(Q(publie_au_raa=False) | Q(publie_au_raa__isnull=True))

    if demandeur:
        exists_demandeur = (
            Instructeur.objects.annotate(
                nom_complet=Concat(
                    F("id_agent_autorisations__prenom"),
                    Value(" "),
                    F("id_agent_autorisations__nom"),
                )
            )
            .filter(nom_complet__iexact=demandeur)
            .exists()
        )

        if exists_demandeur:
            avis_list = avis_list.annotate(
                nom_complet_demandeur=Concat(
                    Coalesce(F("id_instructeur__id_agent_autorisations__prenom"), Value("")),
                    Value(" "),
                    Coalesce(F("id_instructeur__id_agent_autorisations__nom"), Value("")),
                )
            ).filter(nom_complet_demandeur__iexact=demandeur)
        else:
            demandeur = ""


    if expert:
        # Vérifier existence
        exists_expert = (
            Expert.objects.annotate(
                interne=Concat(
                    Coalesce(F("id_instructeur__id_agent_autorisations__prenom"), Value("")),
                    Value(" "),
                    Coalesce(F("id_instructeur__id_agent_autorisations__nom"), Value(""))
                ),
                externe=Concat(
                    Coalesce(F("id_contact_externe__prenom"), Value("")),
                    Value(" "),
                    Coalesce(F("id_contact_externe__nom"), Value(""))
                )
            )
            .filter(Q(interne__iexact=expert) | Q(externe__iexact=expert))
            .exists()
        )

        if exists_expert:
            avis_list = avis_list.annotate(
                nom_complet_expert=Concat(
                    Coalesce(F("id_expert__id_instructeur__id_agent_autorisations__prenom"), Value("")),
                    Value(" "),
                    Coalesce(F("id_expert__id_instructeur__id_agent_autorisations__nom"), Value(""))
                ),
                nom_complet_externe=Concat(
                    Coalesce(F("id_expert__id_contact_externe__prenom"), Value("")),
                    Value(" "),
                    Coalesce(F("id_expert__id_contact_externe__nom"), Value(""))
                )
            ).filter(
                Q(nom_complet_expert__iexact=expert) |
                Q(nom_complet_externe__iexact=expert)
            )
        else:
            expert = ""


    if num_dossier:
        avis_list = avis_list.filter(
            Q(id_dossier__numero=num_dossier)
            | Q(dossieravis__id_dossier__numero=num_dossier)
        ).distinct()


    if type_demarche:
        avis_list = avis_list.filter(id_dossier__id_demarche__type__icontains=type_demarche)

    avis_list = avis_list.order_by("-date_demande_avis")

    # Export Excel
    if request.GET.get("export") == "xlsx":
        return _export_avis_xlsx(avis_list)


    # --- Contexte pour le template ---
    context = {
        "avis_list": avis_list,
        "recherche_avis_effectuee": bool(request.GET),
        "recherche_dossier_effectuee": False,
        "annees_avis": annees_avis,
        "mois_list": mois_list,
        "annees": annees,
        'etapes_dossier': etapes_dossier,
        "groupes": groupes,
        "mois_list": mois_list,
        "demarches": demarches,

        # Champs nettoyés pour pré-remplissage propre
        "expert_rempli": expert,
        "demandeur_rempli": demandeur,
        "num_dossier_rempli": num_dossier or "",
        "num_avis_rempli": num_avis or "",
        "date_debut_demande_avis_rempli": date_debut_demande_avis or "",
        "date_fin_demande_avis_rempli": date_fin_demande_avis or "",

    }

 
    return render(request, "instruction/requetes.html", context)



# @login_required
# def autocomplete_numero_dossier(request):
#     query = request.GET.get("term", "").strip()
#     if not query.isdigit():
#         return JsonResponse([], safe=False)

#     suggestions = Dossier.objects.filter(numero__startswith=query) \
#                                  .order_by("numero")[:5]

#     resultats = [{"value": d.numero} for d in suggestions]
#     return JsonResponse(resultats, safe=False)


@login_required
def autocomplete_numero_dossier(request):
    query = request.GET.get("term", "").strip()
    if not query.isdigit():
        return JsonResponse([], safe=False)

    results = []
    seen = set()

    def add_value(value):
        value = str(value).strip()
        if not value or value in seen:
            return
        seen.add(value)
        results.append({"value": value})

    # Dossiers classiques
    suggestions_dossiers = (
        Dossier.objects
        .filter(numero__startswith=query)
        .order_by("numero")
        .values_list("numero", flat=True)[:5]
    )
    for numero in suggestions_dossiers:
        add_value(numero)

    # Dossiers DM non liés
    dossiers_deja_lies_ids = DossierManifestationLiaison.objects.values_list(
        "id_dossier_manif_id", flat=True
    )

    suggestions_dm = (
        DossierManifSportive.objects
        .exclude(id__in=dossiers_deja_lies_ids)
        .filter(numero_dossier_declaration_manifestations__startswith=query)
        .order_by("numero_dossier_declaration_manifestations")
        .values_list("numero_dossier_declaration_manifestations", flat=True)[:5]
    )
    for numero in suggestions_dm:
        add_value(numero)

    # Tri numérique propre
    results = sorted(results, key=lambda x: int(x["value"]))

    return JsonResponse(results[:5], safe=False)


# @login_required
# def autocomplete_nom_beneficiaire(request):
#     query = request.GET.get("term", "").strip()
#     if not query:
#         return JsonResponse([], safe=False)

#     beneficiaires = ContactExterne.objects.filter(
#         dossierbeneficiaire__isnull=False
#     ).filter(
#         Q(raison_sociale__icontains=query) |
#         Q(nom__icontains=query) |
#         Q(prenom__icontains=query)
#     ).values("raison_sociale", "nom", "prenom").distinct()

#     results = []
#     for b in beneficiaires:
#         if b["raison_sociale"]:
#             display = b["raison_sociale"]
#         elif b["nom"] and b["prenom"]:
#             display = f"{b['nom']} {b['prenom']}"
#         # elif b["nom"]:
#         #     display = b["nom"]
#         else:
#             continue

#         results.append({"value": display})

#     return JsonResponse(results[:5], safe=False)


@login_required
def autocomplete_nom_beneficiaire(request):
    query = request.GET.get("term", "").strip()
    if not query:
        return JsonResponse([], safe=False)

    results = []
    seen = set()

    def add_value(value):
        if not value:
            return
        value = " ".join(str(value).split()).strip()
        if not value:
            return
        key = value.lower()
        if key in seen:
            return
        seen.add(key)
        results.append({"value": value})

    # ---------------------------------
    # Bénéficiaires classiques
    # ---------------------------------
    beneficiaires = (
        ContactExterne.objects
        .filter(dossierbeneficiaire__isnull=False)
        .filter(
            Q(raison_sociale__icontains=query) |
            Q(nom__icontains=query) |
            Q(prenom__icontains=query)
        )
        .values("raison_sociale", "nom", "prenom")
        .distinct()[:10]
    )

    for b in beneficiaires:
        if b["raison_sociale"]:
            add_value(b["raison_sociale"])
        elif b["nom"] and b["prenom"]:
            add_value(f"{b['nom']} {b['prenom']}")

    # ---------------------------------
    # Dossiers DM non liés
    # ---------------------------------
    dossiers_deja_lies_ids = DossierManifestationLiaison.objects.values_list(
        "id_dossier_manif_id", flat=True
    )

    dossiers_dm = DossierManifSportive.objects.exclude(id__in=dossiers_deja_lies_ids)

    # Structures
    structures = (
        dossiers_dm
        .filter(structure__icontains=query)
        .values_list("structure", flat=True)
        .distinct()[:10]
    )
    for s in structures:
        add_value(s)

    # Nom prénom organisateur
    organisateurs = (
        dossiers_dm
        .filter(
            Q(nom_organisateur__icontains=query) |
            Q(prenom_organisateur__icontains=query)
        )
        .values("nom_organisateur", "prenom_organisateur")
        .distinct()[:10]
    )
    for o in organisateurs:
        nom = (o["nom_organisateur"] or "").strip()
        prenom = (o["prenom_organisateur"] or "").strip()
        if nom and prenom:
            add_value(f"{nom} {prenom}")
        elif nom:
            add_value(nom)
        elif prenom:
            add_value(prenom)

    # Nom dossier DM
    noms_dossier = (
        dossiers_dm
        .filter(nom_dossier__icontains=query)
        .values_list("nom_dossier", flat=True)
        .distinct()[:10]
    )
    for nd in noms_dossier:
        add_value(nd)

    return JsonResponse(results[:5], safe=False)



@login_required
def autocomplete_instructeur(request):
    term = request.GET.get("term", "").strip()
    if not term:
        return JsonResponse([], safe=False)

    instructeurs = Instructeur.objects.filter(
        Q(id_agent_autorisations__nom__icontains=term) |
        Q(id_agent_autorisations__prenom__icontains=term)
    ).annotate(
        nom_complet=Concat(
            F("id_agent_autorisations__nom"),
            Value(" "),
            F("id_agent_autorisations__prenom")
        )
    ).values_list("nom_complet", flat=True).distinct().order_by("nom_complet")[:5]


    data = [{"value": nom} for nom in instructeurs]
    return JsonResponse(data, safe=False)



@login_required
def autocomplete_demandeur(request):
    """
    Retourne la liste des instructeurs (demandeurs d'avis envoyés)
    """
    term = request.GET.get("term", "").strip()

    results = (
        Instructeur.objects.filter(
            avis__statut="Envoyé",  # instructeurs liés à des avis envoyés
        )
        .filter(
            Q(id_agent_autorisations__nom__icontains=term)
            | Q(id_agent_autorisations__prenom__icontains=term)
        )
        .select_related("id_agent_autorisations")
        .distinct()[:15]
    )

    suggestions = [
        {
            "label": f"{i.id_agent_autorisations.prenom or ''} {i.id_agent_autorisations.nom or ''}".strip(),
            "value": f"{i.id_agent_autorisations.prenom or ''} {i.id_agent_autorisations.nom or ''}".strip(),
        }
        for i in results
        if i.id_agent_autorisations
    ]

    return JsonResponse(suggestions, safe=False)



@login_required
def autocomplete_expert(request):
    """
    Retourne la liste des experts (internes et externes) liés à des avis envoyés.
    """
    term = request.GET.get("term", "").strip()

    results = (
        Expert.objects.filter(avis__statut="Envoyé")
        .filter(
            Q(id_instructeur__id_agent_autorisations__nom__icontains=term)
            | Q(id_instructeur__id_agent_autorisations__prenom__icontains=term)
            | Q(id_contact_externe__nom__icontains=term)
            | Q(id_contact_externe__prenom__icontains=term)
        )
        .select_related(
            "id_instructeur__id_agent_autorisations",
            "id_contact_externe"
        )
        .distinct()[:15]
    )

    suggestions = []
    for e in results:
        if e.id_instructeur and e.id_instructeur.id_agent_autorisations:
            a = e.id_instructeur.id_agent_autorisations
            nom = f"{a.prenom or ''} {a.nom or ''}".strip()
        elif e.id_contact_externe:
            c = e.id_contact_externe
            nom = f"{c.prenom or ''} {c.nom or ''}".strip()
        else:
            continue

        suggestions.append({"label": nom, "value": nom})

    return JsonResponse(suggestions, safe=False)

